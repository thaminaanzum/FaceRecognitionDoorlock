import cv2, os, sys, pickle
import time
import numpy as np
from configs import *
import openpyxl
import urllib.request

from datetime import date

from datetime import datetime


faceCascade = cv2.CascadeClassifier('bin/haarcascade_frontalface_default.xml')
profileCascade = cv2.CascadeClassifier('bin/haarcascade_profileface.xml')

recognizer =cv2.face_LBPHFaceRecognizer.create()
recognizer.read('face_rec_saved_model.yaml')

with open(label_name_map_file, 'rb') as handle:
        label_name_map = pickle.load(handle)

print("Press 'q' to quit\n\n\n")
def ece(d):

   
    path = "ece.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active
  
    sheet = wb_obj.active

    # print the total number of rows
    s=sheet.max_row
    if s>0:
        s=s+1
    print(s)
    #here donot declare  c0
    #writing values to cells   
    c1 = sheet.cell(s, column = 1)                    
    c1.value = d[0]
                          
    c2 = sheet.cell(s , column = 2)
    c2.value = d[1]

    c3 = sheet.cell(s , column = 3)
    c3.value = d[2]

    c4 = sheet.cell(s , column = 4)
    c4.value = d[3]

    #c6 = sheet.cell(s , column = 6)      for add column
    #c6.value = d[6]

    wb_obj.save("ece.xlsx")
def predictFacesFromWebcam(label2name_map):
        url = "http://192.168.239.1/cam-mid.jpg"
        cam = cv2.VideoCapture(url)

        d=['0','0','0','0','0']
        """d[0]="jjj"
        today = date.today()
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        d[1]=str(today)
        d[2]=current_time
        d[3]="present"
        ece(d)"""
        while True:
                img_resp=urllib.request.urlopen(url)
                imgnp=np.array(bytearray(img_resp.read()),dtype=np.uint8)
                #ret, frame = cap.read()
                img_rgb= cv2.imdecode(imgnp,-1)

                
                
                gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
                faces = faceCascade.detectMultiScale(gray, scaleFactor, minNeighbors, cascadeFlags, minSize)
                for (x, y, w, h) in faces:
                        cv2.rectangle(img_rgb, (x, y), (x+w, y+h), (0, 255, 0), 2)
                        name_predicted, confidence = recognizer.predict(cv2.resize(gray[y: y + h, x: x + w], face_resolution))
                        print(str(name_predicted) +' , ' +str(confidence))
                        if(name_predicted!=0 and confidence<confidence_threshold):
                                print("It is predicted as "+label2name_map[name_predicted])
                                cv2.putText(img_rgb, label2name_map[name_predicted], (x+3,y+h+20), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,0))
                                d[0]=str(label2name_map[name_predicted])
                                today = date.today()
                                now = datetime.now()
                                current_time = now.strftime("%H:%M:%S")
                                d[1]=str(today)
                                d[2]=current_time
                                d[3]="present"
                                ece(d)
                cv2.imshow('Video', img_rgb)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                        print("\nQuitting")
                        break
        video_capture.release()
        cvcv2.destroyAllWindows()
predictFacesFromWebcam(label_name_map)
