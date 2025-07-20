import cv2, os, sys, time
import numpy as np
from PIL import Image
from configs import *
import cv2, os, sys, pickle
import time
import numpy as np
from configs import *
from tkinter import *
from PIL import Image, ImageTk
from datetime import date
import openpyxl
from datetime import datetime
import urllib.request
faceCascade = cv2.CascadeClassifier('bin/haarcascade_frontalface_default.xml')
profileCascade = cv2.CascadeClassifier('bin/haarcascade_profileface.xml')
recognizer = cv2.face_LBPHFaceRecognizer.create()
recognizer1 = cv2.face_LBPHFaceRecognizer.create()
cv2.CascadeClassifier('bin/haarcascade_profileface.xml')
recognizer.read('face_rec_saved_model.yaml')
recognizer1.read('face_rec_saved_model.yaml')
with open(label_name_map_file, 'rb') as handle:
        label_name_map = pickle.load(handle)

print("Press 'q' to quit\n\n\n")
#can also use createEigenFaceRecognizer() or createFisherFaceRecognizer() or createLBPHFaceRecognizer()
#Read http://docs.opencv.org/2.4/modules/contrib/doc/facerec/facerec_tutorial.html to understand ML behind
url1 = "http://192.168.68.230/cam-mid.jpg"
video_capture= cv2.VideoCapture(url1)


i=0
#video_capture = cv2.VideoCapture(0) #Set the source webcam
video_capture .set(3,640)
video_capture .set(4,480)

def ece(d):

   
    path = "ece.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active
  
    sheet = wb_obj.active

    # print the total number of rows
    s=sheet.max_row
    if s>0:
        s=s+1
    print(s)
    #here donot declare  c0
    #writing values to cells   
    c1 = sheet.cell(s, column = 1)                    
    c1.value = d[0]
                          
    c2 = sheet.cell(s , column = 2)
    c2.value = d[1]

    c3 = sheet.cell(s , column = 3)
    c3.value = d[2]

    c4 = sheet.cell(s , column = 4)
    c4.value = d[3]

    #c6 = sheet.cell(s , column = 6)      for add column
    #c6.value = d[6]

    wb_obj.save("ece.xlsx")

def predictFacesFromWebcam(label2name_map):
        
        

        d=['0','0','0','0','0']
       
        i=0
        while i==0:
                img_resp=urllib.request.urlopen(url1)
                imgnp=np.array(bytearray(img_resp.read()),dtype=np.uint8)
                frame= cv2.imdecode(imgnp,-1)

               

                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
               
                faces1 = faceCascade.detectMultiScale(gray, scaleFactor, minNeighbors, cascadeFlags, minSize)
                
                for (x, y, w, h) in faces1:
                        cv2.rectangle(frame , (x, y), (x+w, y+h), (0, 255, 0), 2)
                        name_predicted, confidence = recognizer.predict(cv2.resize(gray[y: y + h, x: x + w], face_resolution))
                        print(str(name_predicted) +' , ' +str(confidence))
                        if(name_predicted!=0 and confidence<confidence_threshold):
                                print("It is predicted as "+label2name_map[name_predicted])
                                cv2.putText( frame, label2name_map[name_predicted], (x+3,y+h+20), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,0))
                                i=1
                                d[0]=str(label2name_map[name_predicted])
                                today = date.today()
                                now = datetime.now()
                                current_time = now.strftime("%H:%M:%S")
                                d[1]=str(today)
                                d[2]=current_time
                                d[3]="present"
                                ece(d)

                               
                cv2.imshow('Video',  frame)
                
                if cv2.waitKey(1) & 0xFF == ord('q'):
                        print("\nQuitting")
                        break
        video_capture.release()
        
def main():
    predictFacesFromWebcam(label_name_map)

def getFacesAndNames(path):
	image_paths = [os.path.join(path, f) for f in os.listdir(path)  if f.endswith(pic_format)]
	faces = []
	names = []
	count = 0
	label2name_map = {}
	name2label_map = {}
	
	for i in image_paths:
		
		#Convert to grayscale and get as np_array
		img_gray = Image.open(i).convert('L')
		width, height = img_gray.size
		img = np.array(img_gray, 'uint8')
		#cv2.imshow('q',img)
		#Create label for person
		#name =str(os.path.split(i)[1].split(".")[0].replace("subject", ""))
		person_name = os.path.split(i)[1].split(".")[0]
		print(person_name)
		if person_name in name2label_map:
			name = name2label_map[person_name]
		else:
			count += 1
			name2label_map[person_name] = count
			name = count
			label2name_map[count] = person_name
		
		#to detect frontal face
		face = faceCascade.detectMultiScale(img, scaleFactor, minNeighbors, cascadeFlags, minSize)
		#to detect left side face
		#sideface_left = profileCascade.detectMultiScale(img, scaleFactor, minNeighbors, cascadeFlags, minSize)
		#to detect right side face (mirror flip the image and use same cascade)
		#sideface_right = profileCascade.detectMultiScale(np.fliplr(img), scaleFactor, minNeighbors, cascadeFlags, minSize)
		
		#Add all detected faces to the list
		for(x, y, w, h) in face:
			faces.append(cv2.resize(img[y: y + h, x: x + w], face_resolution))
			names.append(name)
			#cv2.imshow("Adding faces to traning set...", img[y: y + h, x: x + w])
			#cv2.waitKey(0)
			print("Frontal Face found in "+i)
		
		'''for(x, y, w, h) in sideface_left:
			faces.append(cv2.resize(img[y: y + h, x: x + w], face_resolution))
			names.append(name)
			#cv2.imshow("Adding faces to traning set...", img[y: y + h, x: x + w])
			#cv2.waitKey(0)
			print("Left Side Face found in "+i)
			
		for(X, y, w, h) in sideface_right:
			x = width-(X+w) #reflip to unmirror
			faces.append(cv2.resize(img[y: y + h, x: x+w], face_resolution))
			names.append(name)
			#cv2.imshow("Adding faces to traning set...", img[y: y + h, x: x + w])
			#cv2.waitKey(0)
			print("Right Side Face found in "+i)'''
		
		
	return faces, np.array(names), label2name_map


def train():
        
        faces, names, label_name_map = getFacesAndNames('images_db') #Setup the facial pictures
        #cv2.destroyAllWindows()

        recognizer.train(faces, names) #Train for facial recognition
        recognizer.write(outfile) #Dump the trained model
        with open(label_name_map_file, 'wb') as handle:
                pickle.dump(label_name_map, handle, protocol=pickle.HIGHEST_PROTOCOL) #Dump the label:name map






	
def photho1():
        username1 = username_verify.get()
        
        if len(username1)> 0:
                name = str(username1)
                neram = str(int(time.time()))
                i=0
                while i<30:
                        img_resp=urllib.request.urlopen(url1)
                        imgnp=np.array(bytearray(img_resp.read()),dtype=np.uint8)
                        frame= cv2.imdecode(imgnp,-1)
                        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        cv2.imshow('Video', gray)

                        if cv2.waitKey(1) & 0xFF == ord('n'):
                                 cv2.imwrite(db_path+"/"+str(name)+"."+neram+str(i)+".png", gray)
                                 print("Saved as "+str(name)+"."+neram+str(i)+".png"+"\n\n")
                                 i+=1
                                 print("Waiting to capture photo......")                

                print("\n\nPROCESS STOPPED......")
                video_capture.release()
                train()


def name():
    global main_screen1
    main_screen1= Toplevel(login_screen)
    main_screen1.geometry("766x708")
    main_screen1.configure(background='#3d5705')
    global username_verify
    global password_verify
 
    username_verify = StringVar()
    password_verify = StringVar()
 
    global username_login_entry
    global password_login_entry
 
    Label(main_screen1, text="entername",font=("Verdana", 15),bg="#c6eb73",fg="black").place(x=320,y=200)
    username_login_entry = Entry(main_screen1,justify=RIGHT, textvariable=username_verify,font=('Verdana',15,'bold')).place(x=250,y=240)
    Button(main_screen1, text="Login", width=30, height=2,bd=5, command = photho1,bg="#a6ed07",activebackground="#c6eb73").place(x=270,y=300)
    
        
def login():
    global login_screen
    login_screen = Tk()
    login_screen.title("Login")
    login_screen.geometry("766x708")
    login_screen.configure(background='#3d5705')
    Label(login_screen, text="Please enter details below to login",bg="#c6eb73", font=("Calibri", 30)).place(x=150,y=10)
   
    
    Button(login_screen, text="PHOTO ", width=30, height=2,bd=5, command = name,bg="#a6ed07",activebackground="#c6eb73").place(x=170,y=400)
    Button(login_screen, text="MAIN  ", width=30, height=2,bd=5, command =main ,bg="#a6ed07",activebackground="#c6eb73").place(x=370,y=400)
   
  
def login_sucess():
   z1=Label(login_screen , text="Face SUCCESSES ", bg="green",fg="black",font=("calibri", 11))
   z1.place(x=350,y=500)
   
    
 
# Designing popup for login invalid password
 
def password_not_recognised():
      
      z1=Label(login_screen , text="INVALID USER ", bg="red",fg="black",font=("calibri", 11))
      z1.place(x=350,y=500)
      
      


 
def main_account_screen():
    login()
   
    login_screen.mainloop()
          
main_account_screen()
        
